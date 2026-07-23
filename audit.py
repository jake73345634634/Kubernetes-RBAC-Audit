#!/usr/bin/env python3
"""Kubernetes RBAC Audit.

Scans exported RBAC objects (roles, cluster roles, and their bindings) for
permissions that are commonly abused for privilege escalation or data access,
then reports which subjects (users, groups, service accounts) actually hold
those permissions through a binding.

Export the objects first, e.g.:

    kubectl get roles --all-namespaces -o json          > roles.json
    kubectl get rolebindings --all-namespaces -o json   > rolebindings.json
    kubectl get clusterroles -o json                    > clusterroles.json
    kubectl get clusterrolebindings -o json             > clusterrolebindings.json
"""

import argparse
import json
import sys

from colorama import init as init_colorama, Fore, Style


# --- severity levels -------------------------------------------------------
CRITICAL = "CRITICAL"
HIGH = "HIGH"
MEDIUM = "MEDIUM"

_SEVERITY_ORDER = {CRITICAL: 0, HIGH: 1, MEDIUM: 2}
_SEVERITY_COLOR = {
    CRITICAL: Fore.RED + Style.BRIGHT,
    HIGH: Fore.RED,
    MEDIUM: Fore.YELLOW,
}

# Built-in / default roles that are expected to be broad and are not flagged.
_DEFAULT_PREFIXES = ("system:", "kubernetes-")
_DEFAULT_NAMES = ("edit", "admin", "cluster-admin", "aws-node")


def parse_args():
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--roles", metavar="FILE", help="Roles JSON file")
    parser.add_argument("--roleBindings", metavar="FILE", help="RoleBindings JSON file")
    parser.add_argument("--clusterRoles", metavar="FILE", help="ClusterRoles JSON file")
    parser.add_argument("--clusterRoleBindings", metavar="FILE", help="ClusterRoleBindings JSON file")
    parser.add_argument("--output", metavar="FILE",
                        help="Also write pentest-ready reports: FILE-exposed.md and "
                             "FILE-unbound.md (concise 'Affects' tables) plus FILE.xlsx "
                             "(one 'Evidence' workbook with Exposed Roles and Unbound Roles "
                             "tabs). A trailing .md/.xlsx extension on FILE is ignored.")
    return parser.parse_args()


def _fatal(message):
    print(f"{Fore.RED}[error]{Style.RESET_ALL} {message}", file=sys.stderr)
    sys.exit(2)


def load_items(file_path):
    """Load a `kubectl get ... -o json` file and return its `items` list.

    Fails clearly (exit code 2) for any missing file, bad encoding, invalid
    JSON, or output that is not a Kubernetes list object.
    """
    try:
        with open(file_path, encoding="utf-8") as handle:
            data = json.load(handle)
    except FileNotFoundError:
        _fatal(f"file not found: {file_path}")
    except IsADirectoryError:
        _fatal(f"expected a file but got a directory: {file_path}")
    except UnicodeDecodeError:
        _fatal(f"{file_path} is not UTF-8 encoded. Re-export or re-encode it as UTF-8.")
    except json.JSONDecodeError as err:
        _fatal(f"{file_path} is not valid JSON ({err}).")

    if not isinstance(data, dict) or "items" not in data:
        _fatal(f"{file_path} does not look like `kubectl get ... -o json` output "
               "(missing top-level 'items').")

    items = data["items"]
    if items is None:
        return []
    if not isinstance(items, list):
        _fatal(f"{file_path}: 'items' is not a list.")
    return items


class RbacAuditor:
    """Collects risky roles and the subjects bound to them."""

    def __init__(self):
        # key = (kind, namespace, name) -> role record
        self._roles = {}
        self._bindings_loaded = False

    # -- scanning -----------------------------------------------------------
    def scan_roles(self, items, kind):
        for entity in items:
            metadata = entity.get("metadata") or {}
            name = metadata.get("name")
            if not name or self._is_default(name):
                continue

            namespace = metadata.get("namespace") if kind == "Role" else None
            rules = entity.get("rules")  # may be absent (e.g. aggregated roles) or null
            if not rules:
                continue

            for rule in rules:
                resources = rule.get("resources") or []
                verbs = rule.get("verbs") or []
                if not resources or not verbs:
                    continue
                for severity, issue in self._evaluate_rule(resources, verbs):
                    self._add_finding(kind, namespace, name, severity, issue)

    @staticmethod
    def _evaluate_rule(resources, verbs):
        """Return a list of (severity, issue) for a single policy rule."""
        findings = []
        res = set(resources)
        vrb = set(verbs)
        read_verbs = {"*", "get", "list"}

        # Full cluster-admin: every verb on every resource.
        if "*" in res and "*" in vrb:
            findings.append((CRITICAL, "full admin"))

        # A specific dangerous verb against every resource ('*').
        if "*" in res:
            for verb in ("delete", "deletecollection", "create", "impersonate", "list", "get"):
                if verb in vrb:
                    sev = HIGH if verb in {"delete", "deletecollection", "create", "impersonate"} else MEDIUM
                    findings.append((sev, f"{verb} any resource"))
                    break

        # Any verb ('*') on a security-sensitive resource.
        sensitive = ("secrets", "configmaps", "pods", "deployments", "daemonsets",
                     "statefulsets", "replicationcontrollers", "replicasets", "cronjobs",
                     "jobs", "roles", "clusterroles", "rolebindings", "clusterrolebindings",
                     "users", "groups")
        if "*" in vrb:
            hit = next((r for r in sensitive if r in res), None)
            if hit:
                findings.append((HIGH, f"perform any verb on {hit}"))

        # Read access to secrets / configmaps.
        if "secrets" in res and (vrb & read_verbs):
            findings.append((HIGH, "read secrets"))
        if "configmaps" in res and (vrb & read_verbs):
            findings.append((MEDIUM, "read configmaps"))

        # Privilege escalation via creating roles or bindings.
        for r in ("clusterrolebindings", "rolebindings", "clusterroles", "roles"):
            if r in res and ("create" in vrb or "*" in vrb):
                findings.append((HIGH, f"create {r}"))
                break

        # Creating/updating pod-spawning workloads.
        pod_spawning = ("pods", "deployments", "daemonsets", "statefulsets",
                        "replicationcontrollers", "replicasets", "jobs", "cronjobs")
        spawn_hit = next((r for r in pod_spawning if r in res), None)
        if spawn_hit:
            if "create" in vrb:
                findings.append((HIGH, f"create {spawn_hit}"))
            elif "update" in vrb:
                findings.append((MEDIUM, f"update {spawn_hit}"))

        # Pod subresources used for shell access into running workloads.
        if "pods/exec" in res and (vrb & {"*", "create", "get"}):
            findings.append((HIGH, "exec into pods"))
        if "pods/attach" in res and (vrb & {"*", "create", "get"}):
            findings.append((HIGH, "attach to running pods"))

        return findings

    def _add_finding(self, kind, namespace, name, severity, issue):
        key = (kind, namespace, name)
        role = self._roles.get(key)
        if role is None:
            role = {"kind": kind, "namespace": namespace, "name": name,
                    "findings": [], "subjects": []}
            self._roles[key] = role
        if (severity, issue) not in role["findings"]:
            role["findings"].append((severity, issue))

    # -- bindings -----------------------------------------------------------
    def attach_bindings(self, items, binding_kind):
        self._bindings_loaded = True
        for entity in items:
            metadata = entity.get("metadata") or {}
            binding_name = metadata.get("name", "<unknown>")
            binding_ns = metadata.get("namespace")
            ref = entity.get("roleRef") or {}
            ref_kind = ref.get("kind")
            ref_name = ref.get("name")
            subjects = entity.get("subjects") or []
            if not ref_name or not subjects:
                continue

            # A RoleBinding may reference a Role in its own namespace or a
            # cluster-wide ClusterRole; a ClusterRoleBinding only the latter.
            if ref_kind == "ClusterRole":
                role = self._roles.get(("ClusterRole", None, ref_name))
            elif ref_kind == "Role":
                role = self._roles.get(("Role", binding_ns, ref_name))
            else:
                role = None
            if role is None:
                continue

            for sub in subjects:
                sub_name = sub.get("name")
                if not sub_name:
                    continue
                record = {
                    "kind": sub.get("kind", "?"),
                    "name": sub_name,
                    "namespace": sub.get("namespace"),
                    "binding_kind": binding_kind,
                    "binding_name": binding_name,
                }
                if record not in role["subjects"]:
                    role["subjects"].append(record)

    @staticmethod
    def _is_default(name):
        return (name.startswith(_DEFAULT_PREFIXES) or name in _DEFAULT_NAMES)

    # -- reporting ----------------------------------------------------------
    def report(self):
        roles = list(self._roles.values())

        print(f"{Style.BRIGHT}{'=' * 64}{Style.RESET_ALL}")
        print(f"{Style.BRIGHT} Kubernetes RBAC Audit - findings{Style.RESET_ALL}")
        print(f"{Style.BRIGHT}{'=' * 64}{Style.RESET_ALL}")

        if not roles:
            print(f"{Fore.GREEN}[ok]{Style.RESET_ALL} No risky roles found.")
            return 0

        exposed = [r for r in roles if r["subjects"]]
        sev_counts = {CRITICAL: 0, HIGH: 0, MEDIUM: 0}
        for role in roles:
            sev_counts[self._worst(role)] += 1

        print(f"{len(roles)} risky role(s): "
              f"{_SEVERITY_COLOR[CRITICAL]}{sev_counts[CRITICAL]} critical{Style.RESET_ALL}, "
              f"{_SEVERITY_COLOR[HIGH]}{sev_counts[HIGH]} high{Style.RESET_ALL}, "
              f"{_SEVERITY_COLOR[MEDIUM]}{sev_counts[MEDIUM]} medium{Style.RESET_ALL}.")

        if self._bindings_loaded:
            print(f"{Fore.RED}{Style.BRIGHT}{len(exposed)}{Style.RESET_ALL} of them are "
                  f"{Fore.RED}{Style.BRIGHT}EXPOSED{Style.RESET_ALL} (bound to a subject) - fix these first.")
        else:
            print(f"{Fore.YELLOW}[note]{Style.RESET_ALL} No binding files supplied; "
                  "cannot tell which risky roles are actually granted to anyone.")

        # Exposed + worst-severity first so the biggest problems are at the top.
        for role in sorted(roles, key=self._sort_key):
            self._print_role(role)
        print()
        # Non-zero exit lets CI/pipelines fail the build on findings.
        return 1

    def _sort_key(self, role):
        exposed_first = 0 if role["subjects"] else 1
        return (exposed_first, _SEVERITY_ORDER[self._worst(role)], role["kind"], role["name"])

    @staticmethod
    def _worst(role):
        return min((sev for sev, _ in role["findings"]), key=lambda s: _SEVERITY_ORDER[s])

    def _print_role(self, role):
        location = f"{role['kind']} "
        location += f"{role['namespace']}/{role['name']}" if role["namespace"] else role["name"]

        if role["subjects"]:
            tag = f"{Fore.RED}{Style.BRIGHT}[EXPOSED]{Style.RESET_ALL}"
        elif self._bindings_loaded:
            tag = f"{Fore.YELLOW}[unbound]{Style.RESET_ALL}"
        else:
            tag = f"{Fore.YELLOW}[risky]{Style.RESET_ALL}"

        print()
        print(f"{tag} {Style.BRIGHT}{location}{Style.RESET_ALL}")
        for sev, issue in sorted(role["findings"], key=lambda f: _SEVERITY_ORDER[f[0]]):
            print(f"    {_SEVERITY_COLOR[sev]}{sev:<8}{Style.RESET_ALL} {issue}")
        for sub in role["subjects"]:
            who = f"{sub['namespace']}/{sub['name']}" if sub["namespace"] else sub["name"]
            print(f"      {Fore.CYAN}-> granted to{Style.RESET_ALL} {sub['kind']}: {who} "
                  f"(via {sub['binding_kind']} {sub['binding_name']})")

    # -- report helpers -----------------------------------------------------
    EXPOSED_TITLE = "Exposed RBAC Roles"
    UNBOUND_TITLE = "Unbound RBAC Roles"

    @staticmethod
    def _location(role):
        if role["namespace"]:
            return f"{role['kind']} {role['namespace']}/{role['name']}"
        return f"{role['kind']} {role['name']}"

    @staticmethod
    def _subject_who(sub):
        return f"{sub['namespace']}/{sub['name']}" if sub["namespace"] else sub["name"]

    def _exposed_roles(self):
        return [r for r in sorted(self._roles.values(), key=self._sort_key) if r["subjects"]]

    def _unbound_roles(self):
        return [r for r in sorted(self._roles.values(), key=self._sort_key) if not r["subjects"]]

    def _access_bullets(self, role):
        """Effective-access issues (no severities) as a list of bullet lines."""
        items = sorted(role["findings"], key=lambda f: _SEVERITY_ORDER[f[0]])
        return [f"- {issue}" for _, issue in items]

    @staticmethod
    def _grid_table(headers, rows):
        """Render a Pandoc grid table.

        `rows` is a list of rows; each row is a list of cells; each cell is a
        list of text lines (so a cell can hold a multi-line bullet list).
        """
        ncols = len(headers)
        widths = [len(h) for h in headers]
        for row in rows:
            for i, cell in enumerate(row):
                for line in cell:
                    widths[i] = max(widths[i], len(line))

        def rule(fill):
            return "+" + "+".join(fill * (w + 2) for w in widths) + "+"

        def render(cells):
            height = max((len(c) for c in cells), default=1)
            lines = []
            for row_line in range(height):
                parts = []
                for i in range(ncols):
                    text = cells[i][row_line] if row_line < len(cells[i]) else ""
                    parts.append(" " + text.ljust(widths[i]) + " ")
                lines.append("|" + "|".join(parts) + "|")
            return lines

        out = [rule("-")]
        out += render([[h] for h in headers])
        out.append(rule("="))
        for row in rows:
            out += render(row)
            out.append(rule("-"))
        return "\n".join(out)

    def _summary_line(self, roles):
        counts = {CRITICAL: 0, HIGH: 0, MEDIUM: 0}
        for role in roles:
            counts[self._worst(role)] += 1
        return (f"**{len(roles)} role(s):** {counts[CRITICAL]} critical, "
                f"{counts[HIGH]} high, {counts[MEDIUM]} medium.")

    # -- markdown reports ---------------------------------------------------
    def markdown_exposed(self):
        roles = self._exposed_roles()
        out = [f"# {self.EXPOSED_TITLE}", ""]
        if not roles:
            out.append("None: no risky role is currently bound to a subject.")
            return "\n".join(out) + "\n"

        out += [self._summary_line(roles), ""]

        # A subject is often bound to several different roles that grant the
        # same effective access (e.g. many cert-manager roles all "read
        # secrets"). Collapse those into one row per (subject, access) and show
        # a "Grants" count of how many distinct roles produce it, so identical
        # rows don't repeat.
        groups = {}
        order = []
        for role in roles:
            access = tuple(self._access_bullets(role))
            role_key = (role["kind"], role["namespace"], role["name"])
            for sub in role["subjects"]:
                subject = f"{sub['kind']}: {self._subject_who(sub)}"
                key = (subject, access)
                group = groups.get(key)
                if group is None:
                    group = {"severity": self._worst(role), "subject": subject,
                             "access": access, "roles": set()}
                    groups[key] = group
                    order.append(key)
                group["roles"].add(role_key)

        rows = [[[g["severity"]], [g["subject"]], [str(len(g["roles"]))], list(g["access"])]
                for g in (groups[k] for k in order)]
        out.append(self._grid_table(
            ["Severity", "Granted To", "Grants", "Effective Access"], rows))
        return "\n".join(out) + "\n"

    def markdown_unbound(self):
        roles = self._unbound_roles()
        out = [f"# {self.UNBOUND_TITLE}", ""]
        if not roles:
            out.append("None: every risky role is bound to a subject.")
            return "\n".join(out) + "\n"

        out += [self._summary_line(roles), ""]
        rows = [[[self._worst(role)], [self._location(role)], self._access_bullets(role)]
                for role in roles]
        out.append(self._grid_table(["Severity", "Role", "Effective Access"], rows))
        return "\n".join(out) + "\n"

    # -- xlsx report (single workbook, two tabs) ----------------------------
    _EXPOSED_HEADERS = ["Severity", "Kind", "Namespace", "Role", "Effective Access",
                        "Subject Kind", "Subject Namespace", "Subject",
                        "Binding Kind", "Binding"]
    _UNBOUND_HEADERS = ["Severity", "Kind", "Namespace", "Role", "Effective Access"]

    def _exposed_rows(self):
        """One row per permission per grant; every column populated."""
        for role in self._exposed_roles():
            namespace = role["namespace"] or ""
            for sev, issue in sorted(role["findings"], key=lambda f: _SEVERITY_ORDER[f[0]]):
                for sub in role["subjects"]:
                    yield [sev, role["kind"], namespace, role["name"], issue,
                           sub["kind"], sub["namespace"] or "", sub["name"],
                           sub["binding_kind"], sub["binding_name"]]

    def _unbound_rows(self):
        """One row per risky permission of each unbound role."""
        for role in self._unbound_roles():
            namespace = role["namespace"] or ""
            for sev, issue in sorted(role["findings"], key=lambda f: _SEVERITY_ORDER[f[0]]):
                yield [sev, role["kind"], namespace, role["name"], issue]

    @staticmethod
    def _add_table_sheet(ws, table_name, headers, rows):
        """Fill a sheet with a native Excel table, no other formatting.

        Styled 'Dark Teal, Table Style Medium 9' (built-in TableStyleMedium9).
        """
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter

        ws.append(headers)
        for row in rows:
            ws.append(row)
        ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9",
                                              showRowStripes=True)
        ws.add_table(table)

    def write_xlsx(self, file_path):
        """Single workbook: an 'Exposed Roles' tab and an 'Unbound Roles' tab,
        each a native Excel table styled 'Dark Teal, Table Style Medium 9'."""
        from openpyxl import Workbook

        wb = Workbook()
        exposed = wb.active
        exposed.title = "Exposed Roles"
        self._add_table_sheet(exposed, "ExposedRoles",
                              self._EXPOSED_HEADERS, self._exposed_rows())
        unbound = wb.create_sheet("Unbound Roles")
        self._add_table_sheet(unbound, "UnboundRoles",
                              self._UNBOUND_HEADERS, self._unbound_rows())
        wb.save(file_path)


def main():
    args = parse_args()
    init_colorama()

    if not (args.roles or args.clusterRoles):
        _fatal("provide at least one of --roles or --clusterRoles "
               "(bindings alone have nothing to check against).")

    auditor = RbacAuditor()

    # Roles are scanned before bindings so bindings can be matched to them.
    if args.clusterRoles:
        auditor.scan_roles(load_items(args.clusterRoles), "ClusterRole")
    if args.roles:
        auditor.scan_roles(load_items(args.roles), "Role")
    if args.clusterRoleBindings:
        auditor.attach_bindings(load_items(args.clusterRoleBindings), "ClusterRoleBinding")
    if args.roleBindings:
        auditor.attach_bindings(load_items(args.roleBindings), "RoleBinding")

    exit_code = auditor.report()

    if args.output:
        base = args.output
        for ext in (".md", ".xlsx"):
            if base.lower().endswith(ext):
                base = base[: -len(ext)]
                break

        # One Markdown ("Affects") per issue type, plus a single XLSX
        # ("Evidence") workbook with an Exposed and an Unbound tab.
        outputs = [
            (f"{base}-exposed.md", auditor.markdown_exposed, "text"),
            (f"{base}-unbound.md", auditor.markdown_unbound, "text"),
            (f"{base}.xlsx", auditor.write_xlsx, "file"),
        ]
        print()
        for path, producer, kind in outputs:
            try:
                if kind == "text":
                    with open(path, "w", encoding="utf-8") as handle:
                        handle.write(producer())
                else:
                    producer(path)
            except OSError as err:
                _fatal(f"could not write report to {path}: {err}")
            print(f"{Fore.CYAN}[output]{Style.RESET_ALL} wrote {path}")

    return exit_code


if __name__ == "__main__":
    sys.exit(main())
